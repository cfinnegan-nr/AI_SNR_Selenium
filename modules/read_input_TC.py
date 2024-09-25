import openpyxl

def ReadTheInput():
    wrkbk = openpyxl.load_workbook("TestSteps.xlsx")
    sh = wrkbk.active

    # iterate through excel and display data
    mylist = []
    for i in range(1, sh.max_row + 1):
        for j in range(1, sh.max_column + 1):
            cell_obj = sh.cell(row=i, column=j)
            mylist.append(cell_obj.value)
    
    return mylist